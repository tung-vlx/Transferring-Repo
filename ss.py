import selenium
from selenium import webdriver
chromeOpts = webdriver.EdgeOptions()
chromeOpts.add_argument("window-size=960,1080")
chrome = webdriver.Edge(options = chromeOpts)

print("----- Start -----")
def screenshot(stockCode, ws_input, rowIndex):
    chrome.get("http://en.stockbiz.vn/Stocks/" + stockCode + "/Snapshot.aspx")

    from PIL import Image
    def cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y):
        img = Image.open(imageName)
        img = img.crop((elementStart_x, elementStart_y, elementEnd_x, elementEnd_y))
        img.save(imageName)

    from selenium.webdriver.common.by import By
    elementStart = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_TopZone"]/tbody/tr/td/table/tbody/tr[1]/td/table/tbody/tr/td/div')
    elementStart_x = elementStart.location['x']
    elementStart_y = elementStart.location['y']
    elementEnd = chrome.find_element(By.XPATH, '//*[@id="ctl00_webPartManager_wp839831864_wp747632477_cbOfficers"]/div')
    elementEnd_x = elementEnd.location['x'] + elementEnd.size['width'] + 10
    elementEnd_y = elementEnd.location['y'] + elementEnd.size['height']
    imageName = "Excel/resource/"+ stockCode + "-1.png"
    chrome.save_screenshot(imageName)
    cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y)

    element_EnglishName = chrome.find_element(By.CLASS_NAME, 'CompanyTitle').text
    # EnglishName = element_EnglishName[0:element_EnglishName.rindex('(')-1]
    # ws_input.cell(row = rowIndex, column = 4).value = EnglishName
    Exchange = element_EnglishName[element_EnglishName.rindex(':')+2:element_EnglishName.rindex(')')]
    ws_input.cell(row = rowIndex, column = 6).value = Exchange
    element_BusinessDesciption = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[1]').text
    if element_BusinessDesciption == '':
        element_BusinessDesciption = '-'
    ws_input.cell(row = rowIndex, column = 7).value = element_BusinessDesciption
    try:
        element_Website = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr/td[2]/div/table/tbody/tr[3]/td[2]/a').text
    except:
        element_Website = '-'
    if element_Website == '':
        element_Website = '-'
    ws_input.cell(row = rowIndex, column = 9).value = element_Website

    chrome.get("http://en.stockbiz.vn/Stocks/" + stockCode + "/MajorHolders.aspx")
    elementEnd = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr[4]')
    elementEnd_x = elementEnd.location['x'] + elementEnd.size['width'] + 10
    elementEnd_y = elementEnd.location['y'] + elementEnd.size['height']
    imageName = "Excel/resource/"+ stockCode + "-2.png"

    element_Shareholder1 = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr[2]/td[1]').text
    element_Shareholder2 = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr[2]/td[4]').text
    ws_input.cell(row = rowIndex, column = 8).value = element_Shareholder1 + " - " + element_Shareholder2
    chrome.save_screenshot(imageName)
    cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y)

def insertImage(ws, stockCode,rowIndex):
    img = xlsx.drawing.image.Image('Excel/resource/' + stockCode + '-1.png')
    img.anchor = 'A' + str(rowIndex)
    ws.add_image(img)
    img = xlsx.drawing.image.Image('Excel/resource/' + stockCode + '-2.png')
    img.anchor = 'J' + str(rowIndex)
    ws.add_image(img)
    wb.save('screenshot.xlsx')

import openpyxl as xlsx
path = r'C:\Users\vtung010\Downloads\Stockbiz Scraping\Python\Python'
wb = xlsx.load_workbook("Excel/screenshot.xlsx")
ws_input = wb.get_sheet_by_name(name = "Sheet1")

import os
import shutil

print(' -----------------')
try:
    os.mkdir(path + "/Excel/resource")
    print("Create temp folder")
except:
    shutil.rmtree(path + "/Excel/resource")
    os.mkdir(path + "/Excel/resource")
    print("Delete existing temp folder")
    print("Create new temp folder")
print(' -----------------')
rowIndex = 9 # row in Input sheet
while ws_input.cell(row = rowIndex, column = 2).value != None:
    cell_ICB = ws_input.cell(row = rowIndex, column = 2).value
    print(cell_ICB)
    wb.create_sheet("Code " + str(cell_ICB[0:4]))
    print("Create sheet " + "Code " + str(cell_ICB[0:4]))
    ws_ICB = wb.get_sheet_by_name(name = "Code " + cell_ICB[0:4])
    print("Fill the sheet " + "Code " + str(cell_ICB[0:4]))
    #Create the first row and format it
    ws_ICB.cell(row = 1, column = 1).value = cell_ICB + ' - ' + ws_input.cell(row = rowIndex, column = 3).value
    for cell in ws_ICB["1:1"]:
        cell.font = xlsx.styles.Font(bold=True)
        cell.fill = xlsx.styles.PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    ICBTotal = 1
    while ws_input.cell(row = rowIndex + ICBTotal, column = 5).value != None:
        ICBTotal += 1
    ICBTotal -= 1

    print(' -----------------')
    i = 1
    while ws_input.cell(row = rowIndex + 1, column = 5).value != None:
        rowIndex += 1 # move on next row in Input sheet
        if i == 1:
            temp_int = 2 + 50*(i-1)
        else:
            temp_int = 1 + 50 * (i-1)
        for cell in ws_ICB[str(temp_int) + ":" + str(temp_int)]:
            cell.font = xlsx.styles.Font(bold=True)
            cell.fill = xlsx.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        ws_ICB.cell(row = temp_int, column = 1).value = ws_input.cell(row = rowIndex, column = 5).value
        print("Processing....." + str(i) + "/" + str(ICBTotal), end='\r')
        screenshot(ws_input.cell(row = rowIndex, column = 5).value, ws_input, rowIndex)
        insertImage(ws_ICB, ws_input.cell(row = rowIndex, column = 5).value, temp_int+1)
        i += 1
    rowIndex += 1 # move on next row in Input sheet
    print('\n', '-----------------')
print("Save Excel file")
wb.save('screenshot.xlsx')
wb.close()
print("Delete temp folder")
shutil.rmtree(path + "/Excel/resource")
print("Close Chrome")
chrome.quit()
print("----- End -----")
