#statistics
import time
startTime = time.time()
import datetime
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, Color
path = r'C:\Users\vtung010\Downloads\Stockbiz Scraping\Python\Python'

border_style = Border(left=Side(border_style="thin"),
                      right=Side(border_style="thin"),
                      top=Side(border_style="thin"),
                      bottom=Side(border_style="thin"))

def cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y):
    img = Image.open(imageName)
    img = img.crop((elementStart_x, elementStart_y, elementEnd_x, elementEnd_y))
    img.save(imageName)

def insertImage_to_Excel(ws,imageName, cellIndex):
    img = openpyxl.drawing.image.Image(imageName)
    img.anchor = cellIndex
    ws.add_image(img)

def insertImage(ws, stockCode,rowIndex):
    imageName = 'Excel/resource/' + stockCode + '-1.png'
    cellIndex = 'A' + str(rowIndex)
    insertImage_to_Excel(ws, imageName, cellIndex)

    imageName = 'Excel/resource/' + stockCode + '-2.png'
    cellIndex = 'J' + str(rowIndex)
    insertImage_to_Excel(ws, imageName, cellIndex)

def StockCodeScreenshot(stockCode, ws_input, rowIndex):
    chrome2.get("http://en.stockbiz.vn/Stocks/" + stockCode + "/Snapshot.aspx")
    
    from selenium.webdriver.common.by import By
    elementStart = chrome2.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_TopZone"]/tbody/tr/td/table/tbody/tr[1]/td/table/tbody/tr/td/div')
    elementStart_x = elementStart.location['x']
    elementStart_y = elementStart.location['y']
    elementEnd = chrome2.find_element(By.XPATH, '//*[@id="ctl00_webPartManager_wp839831864_wp747632477_cbOfficers"]/div')
    elementEnd_x = elementEnd.location['x'] + elementEnd.size['width'] + 10
    elementEnd_y = elementEnd.location['y'] + elementEnd.size['height']
    imageName = "Excel/resource/"+ stockCode + "-1.png"
    chrome2.save_screenshot(imageName)
    cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y)

    element_EnglishName = chrome2.find_element(By.CLASS_NAME, 'CompanyTitle').text
    EnglishName = element_EnglishName[0:element_EnglishName.rindex('(')-1]
    ws_input.cell(row = rowIndex, column = 4).value = EnglishName
    Exchange = element_EnglishName[element_EnglishName.rindex(':')+2:element_EnglishName.rindex(')')]
    ws_input.cell(row = rowIndex, column = 6).value = Exchange
    element_BusinessDesciption = chrome2.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[1]').text
    if element_BusinessDesciption == '':
        element_BusinessDesciption = '-'
    ws_input.cell(row = rowIndex, column = 7).value = element_BusinessDesciption
    try:
        element_Website = chrome2.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr/td[2]/div/table/tbody/tr[3]/td[2]/a').text
    except:
        element_Website = '-'
    if element_Website == '':
        element_Website = '-'
    ws_input.cell(row = rowIndex, column = 9).value = element_Website

    chrome2.get("http://en.stockbiz.vn/Stocks/" + stockCode + "/MajorHolders.aspx")
    # chrome2.execute_script("document.body.style.zoom='75%'")
    elementEnd = chrome2.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr[4]')
    elementEnd_x = elementEnd.location['x'] + elementEnd.size['width'] + 10
    elementEnd_y = elementEnd.location['y'] + elementEnd.size['height']
    imageName = "Excel/resource/"+ stockCode + "-2.png"

    element_Shareholder1 = chrome2.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr[2]/td[1]').text
    element_Shareholder2 = chrome2.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_CenterZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]/table/tbody/tr[2]/td[4]').text
    ws_input.cell(row = rowIndex, column = 8).value = element_Shareholder1 + " - " + element_Shareholder2
    chrome2.save_screenshot(imageName)
    cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y)
    
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

print(' -------------------')
print(' ------ Start ------')
print(' -------------------')

import os
import shutil
try:
    os.mkdir(path + "/Excel/resource")
    print("Create temp folder")
except:
    shutil.rmtree(path + "/Excel/resource")
    os.mkdir(path + "/Excel/resource")
    print("Delete existing temp folder")
    print("Create new temp folder")

path = r'C:\Users\vtung010\Downloads\Stockbiz Scraping\Python\Python'
wb = openpyxl.load_workbook("Excel/input.xlsx")
ws_input = wb["Sheet1"]
ws_input.sheet_view.showGridLines = False

print('Create client background information')
# Client's and BMS's information
GeneralInfo = ["Client name", "Province, Vietnam", "Local BMS for manufacture/distribution of"]
for rowIndex, info in enumerate(GeneralInfo, start=3):
    cell = ws_input.cell(row=rowIndex, column = 2)
    cell.value = info
    cell.font = Font(name='Georgia', sz=11, b=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

print(' -------------------')
print('Create Local Search Header')
#Local search header
def format_and_merge(start_row, end_row, start_col, end_col, value, font_color, fill_color):
    ws_input.merge_cells(start_row=start_row, end_row=end_row, start_column=start_col, end_column=end_col)
    cell = ws_input.cell(row=start_row, column=start_col)
    cell.value = value
    cell.alignment = Alignment(vertical='center', horizontal='center')
    cell.font = Font(name='Georgia', sz=11, b=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.border = border_style
# Define formatting parameters
year = str(int(datetime.date.today().strftime("%Y")) - 1)
formats = [
    (7, 7, 7, 8, "Stockbiz", 'FFFFFF', 'BF8F00'),
    (7, 7, 9, 11, "Company's website", 'FFFFFF', 'FF0000'),
    (7, 7, 12, 14, f"Annual Report / Unconsolidated Financial Statement 31/12/{year}", '000000', 'FFC000'),
    (7, 7, 15, 18, "Rejection reason", '000000', '92D050')
]
for format_params in formats:
    format_and_merge(*format_params)

def format(value, columnWidth, columnIndex):
    cell = ws_input.cell(row=8, column=columnIndex)
    cell.value = value
    cell.font = Font(name='Georgia', sz=11, color='FFFFFF')
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.fill = PatternFill('solid', fgColor='C00000')
    cell.border = border_style
    columnLetter = get_column_letter(columnIndex)
    ws_input.column_dimensions[columnLetter].width = columnWidth
formats = [("No.", 5.86), ("Vietnamese Name", 42), ("English Name", 42), ("Listed name", 8.71), ("Listed in", 8.86), ("Business Description", 46.43), ("Biggest shareholder", 37.43), ("Website", 26.57), ("Website obtained from web search", 13), ("Website Description", 44.86), (f"Business description\nFY{year}", 33.71), (f"Business segment\nFY{year}", 33.71), (f"Biggest shareholder\nFY{year}", 33.71), ("Non-independence (25%)", 12), ("Different functions", 12), ("Different products", 12), ("Others", 12), ("Accept/Reject", 12), ("Comments", 36), ("Notes", 22.71)]
ws_input.row_dimensions[8].height = 59.25
i=2
for format_params in formats:
    format(*format_params, i)
    i += 1

# Create Summary Sheet
print(' -------------------')
print('Create sheet Amount of Total Companies')
wb.create_sheet('Amount of Total Companies')
ws_summary = wb['Amount of Total Companies']
ws_summary.sheet_view.showGridLines = False
for rowIndex, info in enumerate(GeneralInfo, start=3):
    cell = ws_summary.cell(row=rowIndex, column = 1)
    cell.value = info
    cell.font = Font(name='Georgia', sz=11, b=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')


import selenium
from selenium import webdriver
chromeOpts = webdriver.EdgeOptions()
chromeOpts.add_argument("window-size=960,1080")
chrome = webdriver.Edge(options = chromeOpts)
chrome2 = webdriver.Edge(options = chromeOpts)
from selenium.webdriver.common.by import By
print(' -------------------')
print('Start filling Local Search')
ICBDict = {
    "0001":"Oil & Gas","0500":"Oil & Gas","0530":"Oil & Gas Producers","0533":"Exploration & Production","0537":"Integrated Oil & Gas","0570":"Oil Equipment, Services & Distribution","0573":"Oil Equipment & Services","0577":"Pipelines","0580":"Alternative Energy","0583":"Renewable Energy Equipment","0587":"Alternative Fuels",
    "1000":"Basic Materials","1300":"Chemicals","1350":"Chemicals","1353":"Commodity Chemicals","1357":"Specialty Chemicals","1700":"Basic Resources","1730":"Forestry & Paper","1733":"Forestry","1737":"Paper","1750":"Industrial Metals & Mining","1753":"Aluminum","1755":"Nonferrous Metals","1757":"Iron & Steel","1770":"Mining","1771":"Coal","1773":"Diamonds & Gemstones","1775":"General Mining","1777":"Gold Mining","1779":"Platinum & Precious Metals",
    "2000":"Industrials","2300":"Construction & Materials","2350":"Construction & Materials","2353":"Building Materials & Fixtures","2357":"Heavy Construction","2700":"Industrial Goods & Services","2710":"Aerospace & Defense","2713":"Aerospace","2717":"Defense","2720":"General Industrials","2723":"Containers & Packaging","2727":"Diversified Industrials","2730":"Electronic & Electrical Equipment","2733":"Electrical Components & Equipment","2737":"Electronic Equipment","2750":"Industrial Engineering","2753":"Commercial Vehicles & Trucks","2757":"Industrial Machinery","2770":"Industrial Transportation","2771":"Delivery Services","2773":"Marine Transportation","2775":"Railroads","2777":"Transportation Services","2779":"Trucking","2790":"Support Services","2791":"Business Support Services","2793":"Business Training & Employment Agencies","2795":"Financial Administration","2797":"Industrial Suppliers","2799":"Waste & Disposal Services",
    "3000":"Consumer Goods","3300":"Automobiles & Parts","3350":"Automobiles & Parts","3353":"Automobiles","3355":"Auto Parts","3357":"Tires","3500":"Food & Beverage","3530":"Beverages","3533":"Brewers","3535":"Distillers & Vintners","3537":"Soft Drinks","3570":"Food Producers","3573":"Farming & Fishing","3577":"Food Products","3700":"Personal & Household Goods","3720":"Household Goods & Home Construction","3722":"Durable Household Products","3724":"Nondurable Household Products","3726":"Furnishings","3728":"Home Construction","3740":"Leisure Goods","3743":"Consumer Electronics","3745":"Recreational Products","3747":"Toys","3760":"Personal Goods","3763":"Clothing & Accessories","3765":"Footwear","3767":"Personal Products","3780":"Tobacco","3785":"Tobacco",
    "4000":"Health Care","4500":"Health Care","4530":"Health Care Equipment & Services","4533":"Health Care Providers","4535":"Medical Equipment","4537":"Medical Supplies","4570":"Pharmaceuticals & Biotechnology","4573":"Biotechnology","4577":"Pharmaceuticals",
    "5000":"Consumer Services","5300":"Retail","5330":"Food & Drug Retailers","5333":"Drug Retailers","5337":"Food Retailers & Wholesalers","5370":"General Retailers","5371":"Apparel Retailers","5373":"Broadline Retailers","5375":"Home Improvement Retailers","5377":"Specialized Consumer Services","5379":"Specialty Retailers","5500":"Media","5550":"Media","5553":"Broadcasting & Entertainment","5555":"Media Agencies","5557":"Publishing","5700":"Travel & Leisure","5750":"Travel & Leisure","5751":"Airlines","5752":"Gambling","5753":"Hotels","5755":"Recreational Services","5757":"Restaurants & Bars","5759":"Travel & Tourism",
    "6000":"Telecommunications","6500":"Telecommunications","6530":"Fixed Line Telecommunications","6535":"Fixed Line Telecommunications","6570":"Mobile Telecommunications","6575":"Mobile Telecommunications",
    "7000":"Utilities","7500":"Utilities","7530":"Electricity","7535":"Conventional Electricity","7537":"Alternative Electricity","7570":"Gas, Water & Multiutilities","7573":"Gas Distribution","7575":"Multiutilities","7577":"Water",
    "8000":"Financials","8300":"Banks","8350":"Banks","8355":"Banks","8500":"Insurance","8530":"Nonlife Insurance","8532":"Full Line Insurance","8534":"Insurance Brokers","8536":"Property & Casualty Insurance","8538":"Reinsurance","8570":"Life Insurance","8575":"Life Insurance","8600":"Real Estate","8630":"Real Estate Investment & Services","8633":"Real Estate Holding & Development","8637":"Real Estate Services","8670":"Real Estate Investment Trusts","8671":"Industrial & Office REITs","8672":"Retail REITs","8673":"Residential REITs","8674":"Diversified REITs","8675":"Specialty REITs","8676":"Mortgage REITs","8677":"Hotel & Lodging REITs","8700":"Financial Services","8770":"Financial Services","8771":"Asset Managers","8773":"Consumer Finance","8775":"Specialty Finance","8777":"Investment Services","8779":"Mortgage Finance","8900":"Equity/Nonequity Investments","8980":"Equity Investment Instruments","8985":"Equity Investment Instruments","8990":"Nonequity Investment Instruments","8995":"Nonequity Investment Instruments",
    "9000":"Technology","9500":"Technology","9530":"Software & Computer Services","9533":"Computer Services","9535":"Internet","9537":"Software","9570":"Technology Hardware & Equipment","9572":"Computer Hardware","9574":"Electronic Office Equipment","9576":"Semiconductors","9578":"Telecommunications Equipment"
}
rowIndex = 9    # Index in sheet Input
rowIndex2 = 1   # Index in sheet ICB
rowIndex3 = 7   # Index in sheet Amount of Total Companies
columnIndex3 = 1

ICBIndex = 1 
while ws_input.cell(row=ICBIndex, column=1).value != None and ws_input.cell(row=ICBIndex, column=1).value != '':
    cell_value = ws_input.cell(row=ICBIndex, column=1).value
    print(' -------------------')
    print('Detecting ICB '+str(ICBIndex)+': '+str(cell_value))
    if cell_value in (None, ''):
        break
    ICBCode = str(cell_value).zfill(4)
    try:
        ICBFull = ICBCode + " - " + ICBDict[ICBCode]
    except KeyError:
        break

    print(' -------------------')
    print('ICB is valid, and is '+ICBCode)   
    print(' -------------------')
    print(ICBFull)


    wb.create_sheet("Code "+str(ICBCode))
    ws_ICB = wb["Code "+ str(ICBCode)]
    ws_ICB.sheet_view.showGridLines = False
    print("Create sheet Code " + str(ICBCode))
    rowIndex2 = 1

    # Formatting | Sheet ICB ...
    ws_ICB.cell(row=1, column=1).value = ICBFull
    for cell in ws_ICB[1]:
        cell.font = Font(name='Georgia', sz=11, b=True , color='000000')
        cell.fill = PatternFill('solid', fgColor='F8CBAD')

    ws_input.cell(row=rowIndex, column=2).value = ICBFull

    # Formatting | Sheet Input ...
    ws_input.row_dimensions[rowIndex].height = 14.25
    for cell2 in ws_input[rowIndex][1:]:
        cell2.font = Font(name='Georgia', sz=11, b=True , color='000000')
        cell2.alignment = Alignment(vertical='bottom', wrap_text=False)
        cell2.fill = PatternFill('solid', fgColor='F8CBAD')
        cell2.border = border_style
    
    # Formatting | Sheet Summary
    ws_summary.cell(row=rowIndex3, column=1).value = ICBFull
    ws_summary.row_dimensions[rowIndex3].height = 20.25
    ws_summary.cell(row=rowIndex3, column=1).font = Font(name='Georgia', sz=11, b=True , color='000000')
    ws_summary.cell(row=rowIndex3, column=1).alignment = Alignment(vertical='bottom', wrap_text=False)
    ws_summary.cell(row=rowIndex3, column=1).fill = PatternFill('solid', fgColor='F8CBAD')
    rowIndex3 += 1
       
    # Navigate ICB code...
    chrome.get("https://www.stockbiz.vn/IndustryOverview.aspx?Code=" + ICBCode)
    time.sleep(2)
    try:
        chrome.find_element(By.XPATH, '/html/body/div[2]/div/div[1]/div/button').click()
    except:
        print()

    chrome.execute_script("window.scrollTo(0,670)")     # Scroll down to Data Table
    tempRow = rowIndex
    print(' -------------------')    
    print("Detecting Data Table")
    whileBln = True
   
    while whileBln==True:
        # Extracting data...
        try:
            element_table = chrome.find_element(By.CLASS_NAME, 'dataTable')
        except:
            whileBln = False
            elements_row = []
            break

        element_table = element_table.find_element(By.TAG_NAME, 'tbody')
        elements_row = element_table.find_elements(By.TAG_NAME, 'tr')

        if len(elements_row) == 1:
            break
        
        print(' -------------------')    
        print('Screenshot Data Table')
        elementStart = chrome.find_element(By.XPATH, '//*[@id="ctl00_PlaceHolderContentArea_LeftZone"]/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/div[2]')
        elementStart_x = elementStart.location['x'] - 12
        elementStart_y = elementStart.location['y'] - 670
        elementEnd = chrome.find_element(By.XPATH, '//*[@id="ctl00_webPartManager_wp572523149_wp829393896_callbackTopSymbols"]/div[2]')
        elementEnd_x = elementEnd.location['x'] + elementEnd.size['width'] + 10
        elementEnd_y = elementEnd.location['y'] + elementEnd.size['height'] - 670 + 10
        imageName = "Excel/resource/" + ICBCode + "-" + str(columnIndex3//10 + 1) + ".png"
        chrome.save_screenshot(imageName)
        cropImage(imageName, elementStart_x, elementStart_y, elementEnd_x, elementEnd_y)
        cellIndex = get_column_letter(columnIndex3) + str(rowIndex3) 
        insertImage_to_Excel(ws_summary,imageName, cellIndex)        
        
        print(' -------------------')    
        print("Listing companies")
        # Writing to Excel | Sheet Inout ...
        for element_row in elements_row[1:]:
            elements_td = element_row.find_elements(By.TAG_NAME, 'td')
            elements_span = element_row.find_elements(By.TAG_NAME, 'span')
            rowIndex += 1
            ws_input.cell(row=rowIndex, column=2).value = str(rowIndex - ICBIndex - 9 + 1)
            ws_input.cell(row=rowIndex, column=5).value = elements_td[0].text[0:3]
            ws_input.cell(row=rowIndex, column=3).value = elements_td[1].text
            ws_input.cell(row=rowIndex, column=15).value = '=if(or('+ get_column_letter(14) + str(rowIndex) + '="",'+ get_column_letter(14) + str(rowIndex) + '="-"),if(value(right(' + get_column_letter(8) + str(rowIndex) + ',6))>=0.25,"X",""),if(value(right('+ get_column_letter(14) + str(rowIndex) +',6))>=0.25,"X",""))'
            ws_input.cell(row=rowIndex, column=19).value = '=if('+ get_column_letter(20) + str(rowIndex) + '="","Accept","Reject")'
            print("# Proccessing "+elements_td[0].text[0:3]+" ...")
            # Formatting | Sheet Input...
            ws_input.row_dimensions[rowIndex].height = 49.50
            for cell2 in ws_input[rowIndex][1:]:
                cell2.font = Font(name='Georgia', sz=11 , color='000000')
                cell2.alignment = Alignment(wrap_text=True, vertical='center')
                cell2.border = border_style
            for column in [2,5,6,15,16,17,18,19]:
                ws_input.cell(row=rowIndex, column=column).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            # Writing to Excel | Sheet ICB ...
            if rowIndex-tempRow == 1:
                rowIndex2 = 2 + 50*(rowIndex-tempRow-1)
            else:
                rowIndex2 = 1 + 50*(rowIndex-tempRow-1)
            ws_ICB.cell(row=rowIndex2, column=1).value = elements_td[0].text[0:3]
            for cell in ws_ICB[rowIndex2]:
                cell.font = Font(name='Georgia', b=True, sz=11 , color='000000')
                cell.fill = PatternFill('solid', fgColor='FFFF00')
            StockCodeScreenshot(elements_td[0].text[0:3], ws_input, rowIndex)
            insertImage(ws_ICB, elements_td[0].text[0:3], rowIndex2+1)

        # Checking for next page...
        elements_next = chrome.find_element(By.CLASS_NAME, 'pageNavigation')
        element_next = elements_next.find_elements(By.TAG_NAME, 'a')
        if len(element_next) == 0:
            whileBln = False
            continue
        elif element_next[len(element_next)-1].text[0:1] == "T":
            print("Move to next page")
            columnIndex3 += 11
            element_next[len(element_next)-1].click()
            whileBln = True
            time.sleep(1)
        else:
            whileBln = False
            continue
    
    columnIndex3 = 1
    imageName = "Excel/resource/" + ICBCode + "-1.png"
    img = Image.open(imageName)
    img_height = img.height
    rowIndex3 += img_height // 20 + 2
    # key number of companies under ICB code
    if len(elements_row) <= 2:
            ws_input.cell(row=tempRow, column=3).value = str(rowIndex - tempRow) + " company"
    else:
        ws_input.cell(row=tempRow, column=3).value = str(rowIndex - tempRow) + " companies"
    
    rowIndex += 1
    ICBIndex += 1

print(' -------------------')    
print("Close browsers")
chrome.quit()
chrome2.quit()
print(' -------------------')    
print("Save file")
wb.save('Local Search & Screenshot.xlsx')
print(' -------------------')    
print("Delete temp folder")
try:
    shutil.rmtree(path + "/Excel/resource")
except:
    print()

print(' -------------------')    
print("End")
print(' -------------------')    
print("Factos:")

processing_time = int(time.time() - startTime)
if processing_time < 60:
    print("# Running Time: " + str(processing_time) + " s")
elif processing_time < 3599:
    print("# Running Time: " + str(processing_time//60) + "m" + str(processing_time%60) + "s")    

print("# Done: " + str(rowIndex-8-ICBIndex) + " companies")
print("# Made by yoyoitsme")
